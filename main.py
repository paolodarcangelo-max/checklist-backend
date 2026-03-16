from fastapi import FastAPI, UploadFile, File, Depends, HTTPException, Query, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import FileResponse
from pydantic import BaseModel
from typing import List, Optional, Any
from datetime import datetime, timedelta
from jose import jwt, JWTError
import os
import uuid
import json
import sqlite3
import requests
import smtplib
from email.message import EmailMessage

from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

app = FastAPI(title="Checklist Backend")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

SECRET_KEY = os.getenv("JWT_SECRET", "dev-secret")
ALGORITHM = "HS256"
ACCESS_TOKEN_MINUTES = 60 * 24
security = HTTPBearer()

USERS = {
    "tecnico1": {
        "username": "tecnico1",
        "password": "password123",
        "role": "tech",
    }
}

def _pick_first(row: dict, keys: list[str]) -> str:
    for key in keys:
        value = row.get(key)
        if value is None:
            continue
        value = str(value).strip()
        if value:
            return value
    return ""


def get_all_syncrogest_clients(token_uid: str):
    url = f"{SYNCROGEST_BASE}/ws_clienti/clienti"

    all_rows = []
    offset = 0
    page_size = 200

    while True:
        payload = {
            "token_uid": token_uid,
            "num": page_size,
            "offset": offset,
            "find": "",
            "only_clients": 1,
        }

        r = requests.post(url, headers=sg_headers(), json=payload, timeout=30)
        r.raise_for_status()
        data = r.json()

        rows = data.get("data", {}).get("clienti", [])
        if not rows:
            break

        all_rows.extend(rows)

        if len(rows) < page_size:
            break

        offset += len(rows)

    return all_rows

def get_clients_lookup(token_uid: str):
    rows = get_all_syncrogest_clients(token_uid)
    lookup = {}

    for row in rows:
        client_id = str(
            row.get("anagrafica_id")
            or row.get("cliente_id")
            or ""
        ).strip()

        client_name = str(
            row.get("anagrafica_ragione_sociale")
            or row.get("cliente_nome")
            or row.get("ragione_sociale")
            or ""
        ).strip()

        if client_id and client_name:
            lookup[client_id] = client_name

    return lookup

def get_all_syncrogest_plants(token_uid: str):
    url = f"{SYNCROGEST_BASE}/ws_impianti/impianti"

    all_rows = []
    offset = 0
    page_size = 200

    while True:
        payload = {
            "token_uid": token_uid,
            "num": page_size,
            "offset": offset,
        }

        r = requests.post(url, headers=sg_headers(), json=payload, timeout=30)
        r.raise_for_status()
        data = r.json()

        rows = data.get("data", {}).get("impianti", [])
        if not rows:
            break

        all_rows.extend(rows)

        if len(rows) < page_size:
            break

        offset += len(rows)

    return all_rows

def normalize_plant_row(row: dict, clients_lookup: dict | None = None):
    client_id = _pick_first(row, [
        "cliente_id",
        "anagrafica_id",
        "client_id",
    ])

    client_name = _pick_first(row, [
        "cliente_nome",
        "ragione_sociale",
        "anagrafica_ragione_sociale",
        "client_name",
    ])

    if not client_name and clients_lookup and client_id in clients_lookup:
        client_name = clients_lookup[client_id]

    plant_id = _pick_first(row, [
        "impianto_id",
        "id",
    ])

    plant_name = _pick_first(row, [
        "impianto_nome",
        "nome",
        "descrizione",
        "impianto_descrizione",
        "plant_name",
    ])

    address = _pick_first(row, [
        "impianto_indirizzo",
        "indirizzo",
        "ubicazione",
        "impianto_ubicazione",
        "address",
    ])

    matricola = _pick_first(row, [
        "impianto_matricola",
        "matricola",
        "codice",
        "impianto_codice",
    ])

    return {
        "client_id": client_id,
        "client_name": client_name,
        "plant_id": plant_id,
        "plant_name": plant_name,
        "address": address,
        "matricola": matricola,
    }

SYNCROGEST_BASE = "https://app.syncrogest.it/api/v1"
SYNCROGEST_API_KEY = os.getenv("SYNCROGEST_API_KEY", "")
SYNCROGEST_TOKEN_UID = os.getenv("SYNCROGEST_TOKEN_UID", "")
SYNCROGEST_USERNAME = os.getenv("SYNCROGEST_USERNAME", "")
SYNCROGEST_PASSWORD = os.getenv("SYNCROGEST_PASSWORD", "")

PUBLIC_BASE_URL = os.getenv("PUBLIC_BASE_URL", "").rstrip("/")

SMTP_HOST = os.getenv("SMTP_HOST", "")
SMTP_PORT = int(os.getenv("SMTP_PORT") or "587")
SMTP_USER = os.getenv("SMTP_USER", "")
SMTP_PASSWORD = os.getenv("SMTP_PASSWORD", "")
SMTP_FROM = os.getenv("SMTP_FROM", "")

UPLOAD_DIR = "./uploads"
PDF_DIR = "./uploads/checklists_pdf"
REPORT_DIR = "./uploads/report_preventivi"
DB_PATH = "./backend_data.db"

os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(PDF_DIR, exist_ok=True)
os.makedirs(REPORT_DIR, exist_ok=True)

def get_all_syncrogest_plants(token_uid: str):
    url = f"{SYNCROGEST_BASE}/ws_impianti/impianti"

    all_rows = []
    offset = 0
    page_size = 500

    while True:
        payload = {
            "token_uid": token_uid,
            "num": page_size,
            "offset": offset,
        }

        r = requests.post(url, headers=sg_headers(), json=payload, timeout=30)
        r.raise_for_status()
        data = r.json()

        rows = data.get("data", {}).get("impianti", [])
        if not rows:
            break

        all_rows.extend(rows)

        if len(rows) < page_size:
            break

        offset += page_size

    return all_rows

def db_conn():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    conn = db_conn()
    cur = conn.cursor()

    cur.execute("""
        CREATE TABLE IF NOT EXISTS uploaded_files (
            file_id TEXT PRIMARY KEY,
            path TEXT NOT NULL,
            created_at TEXT NOT NULL
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS checklists (
            server_id TEXT PRIMARY KEY,
            created_by TEXT NOT NULL,
            created_at TEXT NOT NULL,
            payload_json TEXT NOT NULL,
            has_ko INTEGER NOT NULL,
            has_todo INTEGER NOT NULL,
            has_negative INTEGER NOT NULL,
            quote_candidate INTEGER NOT NULL,
            pdf_created INTEGER NOT NULL DEFAULT 0,
            pdf_path TEXT,
            pdf_url TEXT,
            impianto_pdf_attached INTEGER NOT NULL DEFAULT 0,
            syncrogest_upload_response_json TEXT,
            report_sent INTEGER NOT NULL DEFAULT 0,
            report_sent_at TEXT
        )
    """)

    conn.commit()
    conn.close()


@app.on_event("startup")
def startup_event():
    init_db()


def row_to_checklist_dict(row: sqlite3.Row) -> dict:
    return {
        "server_id": row["server_id"],
        "created_by": row["created_by"],
        "created_at": row["created_at"],
        "payload": json.loads(row["payload_json"]),
        "has_ko": bool(row["has_ko"]),
        "has_todo": bool(row["has_todo"]),
        "has_negative": bool(row["has_negative"]),
        "quote_candidate": bool(row["quote_candidate"]),
        "pdf_created": bool(row["pdf_created"]),
        "pdf_path": row["pdf_path"],
        "pdf_url": row["pdf_url"],
        "impianto_pdf_attached": bool(row["impianto_pdf_attached"]),
        "syncrogest_upload_response": json.loads(row["syncrogest_upload_response_json"]) if row["syncrogest_upload_response_json"] else None,
        "report_sent": bool(row["report_sent"]),
        "report_sent_at": row["report_sent_at"],
    }


def db_insert_file(file_id: str, path: str):
    conn = db_conn()
    conn.execute(
        "INSERT INTO uploaded_files (file_id, path, created_at) VALUES (?, ?, ?)",
        (file_id, path, datetime.utcnow().isoformat()),
    )
    conn.commit()
    conn.close()


def db_get_file_path(file_id: str) -> Optional[str]:
    conn = db_conn()
    row = conn.execute(
        "SELECT path FROM uploaded_files WHERE file_id = ?",
        (file_id,),
    ).fetchone()
    conn.close()
    return row["path"] if row else None


def db_insert_checklist(record: dict):
    conn = db_conn()
    conn.execute("""
        INSERT INTO checklists (
            server_id,
            created_by,
            created_at,
            payload_json,
            has_ko,
            has_todo,
            has_negative,
            quote_candidate,
            pdf_created,
            pdf_path,
            pdf_url,
            impianto_pdf_attached,
            syncrogest_upload_response_json,
            report_sent,
            report_sent_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        record["server_id"],
        record["created_by"],
        record["created_at"],
        json.dumps(record["payload"]),
        int(record["has_ko"]),
        int(record["has_todo"]),
        int(record["has_negative"]),
        int(record["quote_candidate"]),
        int(record["pdf_created"]),
        record["pdf_path"],
        record["pdf_url"],
        int(record["impianto_pdf_attached"]),
        json.dumps(record["syncrogest_upload_response"]) if record["syncrogest_upload_response"] is not None else None,
        int(record["report_sent"]),
        record["report_sent_at"],
    ))
    conn.commit()
    conn.close()


def db_update_checklist_after_pdf(server_id: str, pdf_path: str, pdf_url: str, impianto_pdf_attached: bool, upload_response: Any):
    conn = db_conn()
    conn.execute("""
        UPDATE checklists
        SET pdf_created = 1,
            pdf_path = ?,
            pdf_url = ?,
            impianto_pdf_attached = ?,
            syncrogest_upload_response_json = ?
        WHERE server_id = ?
    """, (
        pdf_path,
        pdf_url,
        int(impianto_pdf_attached),
        json.dumps(upload_response) if upload_response is not None else None,
        server_id,
    ))
    conn.commit()
    conn.close()


def db_get_checklist(server_id: str) -> Optional[dict]:
    conn = db_conn()
    row = conn.execute("SELECT * FROM checklists WHERE server_id = ?", (server_id,)).fetchone()
    conn.close()
    return row_to_checklist_dict(row) if row else None


def db_list_checklists() -> List[dict]:
    conn = db_conn()
    rows = conn.execute("SELECT * FROM checklists ORDER BY created_at DESC").fetchall()
    conn.close()
    return [row_to_checklist_dict(r) for r in rows]


def db_list_pending_quote_reports() -> List[dict]:
    conn = db_conn()
    rows = conn.execute("""
        SELECT * FROM checklists
        WHERE quote_candidate = 1 AND report_sent = 0
        ORDER BY created_at DESC
    """).fetchall()
    conn.close()
    return [row_to_checklist_dict(r) for r in rows]


def db_mark_reports_sent(server_ids: List[str], when_iso: str):
    if not server_ids:
        return
    conn = db_conn()
    conn.executemany(
        "UPDATE checklists SET report_sent = 1, report_sent_at = ? WHERE server_id = ?",
        [(when_iso, sid) for sid in server_ids],
    )
    conn.commit()
    conn.close()


@app.get("/debug/env")
def debug_env():
    return {
        "SYNCROGEST_API_KEY_set": bool(SYNCROGEST_API_KEY),
        "SYNCROGEST_TOKEN_UID_set": bool(SYNCROGEST_TOKEN_UID),
        "SYNCROGEST_USERNAME_set": bool(SYNCROGEST_USERNAME),
        "SYNCROGEST_PASSWORD_set": bool(SYNCROGEST_PASSWORD),
        "PUBLIC_BASE_URL": PUBLIC_BASE_URL,
        "SMTP_HOST_set": bool(SMTP_HOST),
        "SMTP_USER_set": bool(SMTP_USER),
        "SMTP_PASSWORD_set": bool(SMTP_PASSWORD),
        "SMTP_FROM_set": bool(SMTP_FROM),
        "DB_PATH": DB_PATH,
    }


class LoginReq(BaseModel):
    username: str
    password: str


class LoginRes(BaseModel):
    access_token: str
    token_type: str = "bearer"


def create_access_token(sub: str):
    expire = datetime.utcnow() + timedelta(minutes=ACCESS_TOKEN_MINUTES)
    payload = {"sub": sub, "exp": expire}
    return jwt.encode(payload, SECRET_KEY, algorithm=ALGORITHM)


def require_user(creds: HTTPAuthorizationCredentials = Depends(security)):
    token = creds.credentials
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        username = payload.get("sub")
        if username not in USERS:
            raise HTTPException(status_code=401)
        return USERS[username]
    except JWTError:
        raise HTTPException(status_code=401, detail="Invalid token")

@app.get("/syncrogest/debug-plants-pages")
def debug_plants_pages(user=Depends(require_user)):
    token_uid = get_syncrogest_token()
    url = f"{SYNCROGEST_BASE}/ws_impianti/impianti"

    def fetch_page(offset_value: int):
        payload = {
            "token_uid": token_uid,
            "num": 200,
            "offset": offset_value,
        }
        r = requests.post(url, headers=sg_headers(), json=payload, timeout=30)
        r.raise_for_status()
        data = r.json()
        rows = data.get("data", {}).get("impianti", [])
        ids = [
            str(row.get("impianto_id") or row.get("id") or "")
            for row in rows[:10]
        ]
        return {
            "offset": offset_value,
            "count": len(rows),
            "first_ids": ids,
        }

    p0 = fetch_page(0)
    p200 = fetch_page(200)
    p400 = fetch_page(400)

    return {
        "page_0": p0,
        "page_200": p200,
        "page_400": p400,
        "same_0_200": p0["first_ids"] == p200["first_ids"],
        "same_200_400": p200["first_ids"] == p400["first_ids"],
    }

@app.post("/auth/login", response_model=LoginRes)
def login(req: LoginReq):
    user = USERS.get(req.username)
    if not user or req.password != user["password"]:
        raise HTTPException(status_code=401, detail="Bad credentials")
    return LoginRes(access_token=create_access_token(req.username))


def sg_headers():
    if not SYNCROGEST_API_KEY:
        raise HTTPException(status_code=500, detail="SYNCROGEST_API_KEY non configurata")
    return {
        "WS-API-KEY": SYNCROGEST_API_KEY,
        "Content-Type": "application/json",
    }


def get_syncrogest_token():
    if SYNCROGEST_TOKEN_UID:
        return SYNCROGEST_TOKEN_UID

    if not SYNCROGEST_USERNAME or not SYNCROGEST_PASSWORD:
        raise HTTPException(
            status_code=500,
            detail="Manca SYNCROGEST_TOKEN_UID oppure SYNCROGEST_USERNAME/PASSWORD",
        )

    url = f"{SYNCROGEST_BASE}/ws_common/logincheck"
    payload = {
        "username": SYNCROGEST_USERNAME,
        "password": SYNCROGEST_PASSWORD,
    }

    r = requests.post(url, headers=sg_headers(), json=payload, timeout=30)

    try:
        data = r.json()
    except Exception:
        raise HTTPException(status_code=500, detail=f"Risposta Syncrogest non valida: {r.text}")

    token_uid = (
        data.get("token_uid")
        or data.get("data", {}).get("token_uid")
        or data.get("result", {}).get("token_uid")
    )

    if not token_uid:
        raise HTTPException(status_code=500, detail=f"Login Syncrogest fallito: {data}")

    return token_uid

@app.get("/syncrogest/clients")
def get_clients(user=Depends(require_user)):
    token_uid = get_syncrogest_token()
    rows = get_all_syncrogest_clients(token_uid)

    result = []
    for row in rows:
        client_id = str(
            row.get("anagrafica_id")
            or row.get("cliente_id")
            or ""
        ).strip()

        client_name = str(
            row.get("anagrafica_ragione_sociale")
            or row.get("cliente_nome")
            or row.get("ragione_sociale")
            or ""
        ).strip()

        if client_id and client_name:
            result.append({
                "id": client_id,
                "name": client_name,
            })

    unique = {}
    for item in result:
        unique[item["id"]] = item["name"]

    final_result = [
        {"id": k, "name": v}
        for k, v in unique.items()
    ]
    final_result.sort(key=lambda x: x["name"].lower())
    return final_result

@app.get("/syncrogest/plants")
def get_plants(client_id: str = Query(...), user=Depends(require_user)):
    token_uid = get_syncrogest_token()

    url = f"{SYNCROGEST_BASE}/ws_impianti/impianti"
    payload = {
        "token_uid": token_uid,
        "cliente_id": client_id,
        "num": 2000,
        "offset": 0,
    }

    r = requests.post(url, headers=sg_headers(), json=payload, timeout=30)
    r.raise_for_status()
    data = r.json()

    rows = data.get("data", {}).get("impianti", [])
    result = []

    for row in rows:
        plant_name = (
            row.get("impianto_nome")
            or row.get("nome")
            or row.get("descrizione")
            or row.get("impianto_descrizione")
            or ""
        )
        address = (
            row.get("impianto_indirizzo")
            or row.get("indirizzo")
            or row.get("ubicazione")
            or row.get("impianto_ubicazione")
            or ""
        )
        client_name = (
            row.get("cliente_nome")
            or row.get("ragione_sociale")
            or ""
        )
        matricola = (
            row.get("matricola")
            or row.get("impianto_matricola")
            or row.get("codice")
            or row.get("impianto_codice")
            or ""
        )

        result.append({
            "id": str(row.get("impianto_id") or row.get("id") or ""),
            "client_id": client_id,
            "client_name": client_name,
            "address": address,
            "plant_name": plant_name,
            "matricola": matricola,
        })

    return result


@app.get("/syncrogest/plant-by-matricola")
def get_plant_by_matricola(matricola: str = Query(...), user=Depends(require_user)):
    token_uid = get_syncrogest_token()
    rows = get_all_syncrogest_plants(token_uid)

    target = matricola.strip().upper()

    for row in rows:
        row_matricola = str(
            row.get("impianto_matricola")
            or row.get("matricola")
            or row.get("codice")
            or row.get("impianto_codice")
            or ""
        ).strip().upper()

        if row_matricola == target:
            plant_name = (
                row.get("impianto_nome")
                or row.get("nome")
                or row.get("descrizione")
                or row.get("impianto_descrizione")
                or ""
            )

            address = (
                row.get("impianto_indirizzo")
                or row.get("indirizzo")
                or row.get("ubicazione")
                or row.get("impianto_ubicazione")
                or ""
            )

            client_name = (
                row.get("cliente_nome")
                or row.get("ragione_sociale")
                or ""
            )

            return {
                "id": str(row.get("impianto_id") or row.get("id") or ""),
                "client_id": str(row.get("cliente_id") or row.get("anagrafica_id") or ""),
                "client_name": client_name,
                "address": address,
                "plant_name": plant_name,
                "matricola": row_matricola,
            }

    raise HTTPException(status_code=404, detail="Impianto non trovato per matricola")

@app.get("/syncrogest/search-plants")
def search_plants(q: str = Query(...), user=Depends(require_user)):
    token_uid = get_syncrogest_token()

    query = q.strip().lower()
    if not query:
        return []

    rows = get_all_syncrogest_plants(token_uid)
    results = []

    for row in rows:
        plant_name = str(
            row.get("impianto_nome")
            or row.get("nome")
            or row.get("descrizione")
            or row.get("impianto_descrizione")
            or ""
        ).strip()

        address = str(
            row.get("impianto_indirizzo")
            or row.get("indirizzo")
            or row.get("ubicazione")
            or row.get("impianto_ubicazione")
            or ""
        ).strip()

        client_name = str(
            row.get("cliente_nome")
            or row.get("ragione_sociale")
            or ""
        ).strip()

        client_id = str(
            row.get("cliente_id")
            or row.get("anagrafica_id")
            or ""
        ).strip()

        plant_id = str(
            row.get("impianto_id")
            or row.get("id")
            or ""
        ).strip()

        matricola = str(
            row.get("impianto_matricola")
            or row.get("matricola")
            or row.get("codice")
            or row.get("impianto_codice")
            or ""
        ).strip()

        haystack = " | ".join([
            client_name,
            address,
            plant_name,
            matricola,
        ]).lower()

        if query in haystack:
            results.append({
                "id": plant_id,
                "client_id": client_id,
                "client_name": client_name,
                "address": address,
                "plant_name": plant_name,
                "matricola": matricola,
            })

    results.sort(key=lambda x: (
        x["client_name"].lower(),
        x["address"].lower(),
        x["plant_name"].lower(),
    ))

    return results[:100]

@app.get("/syncrogest/debug-clients-count")
def debug_clients_count(user=Depends(require_user)):
    token_uid = get_syncrogest_token()
    rows = get_all_syncrogest_plants(token_uid)

    clients_map = {}
    valid_rows = 0

    for row in rows:
        item = normalize_plant_row(row)
        if item["client_id"] and item["client_name"]:
            valid_rows += 1
            clients_map[item["client_id"]] = item["client_name"]

    sample = sorted(clients_map.values())[:30]

    return {
        "total_plants_rows": len(rows),
        "valid_rows_with_client": valid_rows,
        "unique_clients": len(clients_map),
        "sample_clients": sample,
    }


class CheckItem(BaseModel):
    code: str
    label: str
    ok: bool
    note: Optional[str] = None
    photo_ids: List[str] = []


class TodoItem(BaseModel):
    code: str
    label: str
    selected: bool
    note: Optional[str] = None
    photo_ids: List[str] = []


class ChecklistCreate(BaseModel):
    local_id: str
    client_id: str
    client_name: Optional[str] = None
    plant_id: str
    plant_name: Optional[str] = None
    address: Optional[str] = None
    date_iso: str
    general_notes: Optional[str] = None
    checks: List[CheckItem]
    todos: List[TodoItem]


class QuoteReportSendReq(BaseModel):
    recipients: List[str] = ["info@eadnet.it", "christian@eadnet.it"]
    mark_sent: bool = True


def wrap_text(text: str, max_len: int = 95) -> List[str]:
    if not text:
        return [""]
    words = text.split()
    if not words:
        return [text]

    lines = []
    current = words[0]
    for word in words[1:]:
        if len(current) + 1 + len(word) <= max_len:
            current += " " + word
        else:
            lines.append(current)
            current = word
    lines.append(current)
    return lines


def draw_multiline(c: canvas.Canvas, x: float, y: float, text: str, line_height: float = 5 * mm, max_len: int = 95):
    for line in wrap_text(text, max_len=max_len):
        c.drawString(x, y, line)
        y -= line_height
    return y


def ensure_space(c: canvas.Canvas, y: float, needed_mm: float = 30):
    if y < needed_mm * mm:
        c.showPage()
        return A4[1] - 20 * mm
    return y


def draw_photo_grid(
    c: canvas.Canvas,
    y: float,
    photo_ids: List[str],
    label_x: float = 20 * mm,
    left_x: float = 20 * mm,
    max_width: float = 80 * mm,
    max_height: float = 55 * mm,
    gap_x: float = 8 * mm,
    gap_y: float = 8 * mm,
):
    valid_paths = []
    for photo_id in photo_ids:
        path = db_get_file_path(photo_id)
        if path and os.path.exists(path):
            valid_paths.append(path)

    if not valid_paths:
        return y

    c.setFont("Helvetica-Oblique", 9)
    c.drawString(label_x, y, "Foto:")
    y -= 6 * mm

    images_per_row = 2
    x_positions = [left_x, left_x + max_width + gap_x]
    row_height = max_height + gap_y
    current_index = 0

    while current_index < len(valid_paths):
        y = ensure_space(c, y, needed_mm=75)
        row_paths = valid_paths[current_index:current_index + images_per_row]
        x_idx = 0

        for path in row_paths:
            try:
                img = ImageReader(path)
                iw, ih = img.getSize()
                if iw <= 0 or ih <= 0:
                    x_idx += 1
                    continue

                scale = min(max_width / iw, max_height / ih)
                draw_w = iw * scale
                draw_h = ih * scale

                x = x_positions[x_idx]
                c.drawImage(
                    img,
                    x,
                    y - draw_h,
                    width=draw_w,
                    height=draw_h,
                    preserveAspectRatio=True,
                    mask="auto",
                )
            except Exception:
                pass

            x_idx += 1

        y -= row_height
        current_index += images_per_row

    return y


def generate_checklist_pdf(server_id: str, record: dict) -> str:
    payload = record["payload"]

    file_name = f"checklist_{server_id}.pdf"
    file_path = os.path.join(PDF_DIR, file_name)

    c = canvas.Canvas(file_path, pagesize=A4)
    width, height = A4
    x_left = 15 * mm
    y = height - 15 * mm

    c.setTitle(file_name)

    c.setFont("Helvetica-Bold", 16)
    c.drawString(x_left, y, "Checklist Manutenzione")
    y -= 10 * mm

    c.setFont("Helvetica", 10)
    c.drawString(x_left, y, f"Server ID: {server_id}")
    y -= 6 * mm
    c.drawString(x_left, y, f"Data: {payload.get('date_iso', '')}")
    y -= 6 * mm
    c.drawString(x_left, y, f"Cliente: {payload.get('client_name', '') or ''}")
    y -= 6 * mm
    c.drawString(x_left, y, f"Cliente ID: {payload.get('client_id', '')}")
    y -= 6 * mm
    c.drawString(x_left, y, f"Impianto: {payload.get('plant_name', '') or ''}")
    y -= 6 * mm
    c.drawString(x_left, y, f"Impianto ID: {payload.get('plant_id', '')}")
    y -= 6 * mm
    c.drawString(x_left, y, f"Indirizzo: {payload.get('address', '') or ''}")
    y -= 6 * mm
    c.drawString(x_left, y, f"Tipo impianto: {payload.get('plant_type', '') or ''}")
    y -= 10 * mm

    has_ko = record.get("has_ko", False)
    has_todo = record.get("has_todo", False)
    has_negative = record.get("has_negative", False)

    c.setFont("Helvetica-Bold", 11)
    c.drawString(x_left, y, f"Esito: {'CON ANOMALIE' if has_negative else 'POSITIVO'}")
    y -= 8 * mm
    c.drawString(x_left, y, f"Controlli KO: {'SI' if has_ko else 'NO'}")
    y -= 6 * mm
    c.drawString(x_left, y, f"Lavori da fare: {'SI' if has_todo else 'NO'}")
    y -= 10 * mm

    c.setFont("Helvetica-Bold", 12)
    c.drawString(x_left, y, "CONTROLLI ESEGUITI")
    y -= 7 * mm
    c.setFont("Helvetica", 10)

    for item in payload.get("checks", []):
        y = ensure_space(c, y, needed_mm=40)
        status = "OK" if item.get("ok") else "KO"
        c.drawString(x_left, y, f"- {item.get('label', '')}: {status}")
        y -= 5 * mm

        note = item.get("note") or ""
        if note:
            y = draw_multiline(c, x_left + 5 * mm, y, f"Nota: {note}", max_len=85)
            y -= 2 * mm

        photo_ids = item.get("photo_ids", []) or []
        if photo_ids:
            y = draw_photo_grid(c, y, photo_ids)
            y -= 2 * mm

    y -= 5 * mm
    y = ensure_space(c, y, needed_mm=20)

    c.setFont("Helvetica-Bold", 12)
    c.drawString(x_left, y, "INTERVENTI DA ESEGUIRE")
    y -= 7 * mm
    c.setFont("Helvetica", 10)

    for item in payload.get("todos", []):
        y = ensure_space(c, y, needed_mm=40)
        status = "SI" if item.get("selected") else "NO"
        c.drawString(x_left, y, f"- {item.get('label', '')}: {status}")
        y -= 5 * mm

        note = item.get("note") or ""
        if note:
            y = draw_multiline(c, x_left + 5 * mm, y, f"Nota: {note}", max_len=85)
            y -= 2 * mm

        photo_ids = item.get("photo_ids", []) or []
        if photo_ids:
            y = draw_photo_grid(c, y, photo_ids)
            y -= 2 * mm

    general_notes = payload.get("general_notes") or ""
    if general_notes:
        y = ensure_space(c, y, needed_mm=30)
        y -= 5 * mm
        c.setFont("Helvetica-Bold", 12)
        c.drawString(x_left, y, "NOTE GENERALI")
        y -= 7 * mm
        c.setFont("Helvetica", 10)
        y = draw_multiline(c, x_left, y, general_notes, max_len=95)

    c.save()
    return file_path


def build_public_pdf_url(request: Request, server_id: str) -> str:
    if PUBLIC_BASE_URL:
        return f"{PUBLIC_BASE_URL}/checklists/{server_id}/pdf"
    return str(request.url_for("get_checklist_pdf", server_id=server_id))


def syncrogest_upload_pdf_to_impianto(impianto_id: str, pdf_path: str):
    token_uid = get_syncrogest_token()
    url = f"{SYNCROGEST_BASE}/ws_allegati/upload"

    data = {
        "token_uid": token_uid,
        "module": "MOD_IMPIANTI",
        "ref_id": impianto_id,
    }

    with open(pdf_path, "rb") as f:
        files = {
            "userfile": (os.path.basename(pdf_path), f, "application/pdf")
        }
        r = requests.post(
            url,
            headers={"WS-API-KEY": SYNCROGEST_API_KEY},
            data=data,
            files=files,
            timeout=60
        )

    content_type = r.headers.get("Content-Type", "")
    response_json = r.json() if "application/json" in content_type else {"raw": r.text}

    ok = (
        r.status_code == 200 and
        isinstance(response_json, dict) and
        str(response_json.get("status_code")) == "1"
    )

    return {
        "ok": ok,
        "status_code": r.status_code,
        "json": response_json,
    }


def create_quote_report_excel(items: List[dict]) -> str:
    now_tag = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    file_path = os.path.join(REPORT_DIR, f"report_preventivi_{now_tag}.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "Preventivi"

    headers = [
        "Data checklist",
        "Nome cliente",
        "Cliente ID",
        "Nome impianto",
        "Impianto ID",
        "Indirizzo",
        "Tipo impianto",
        "Esito",
        "Controlli KO",
        "Lavori da fare",
        "Note generali",
        "PDF URL",
        "Server ID",
    ]
    ws.append(headers)

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    for item in items:
        payload = item["payload"]

        ko_labels = [
            c.get("label", "")
            for c in payload.get("checks", [])
            if c.get("ok") is False
        ]

        todo_labels = [
            t.get("label", "")
            for t in payload.get("todos", [])
            if t.get("selected") is True
        ]

        ws.append([
            payload.get("date_iso", ""),
            payload.get("client_name", "") or "",
            payload.get("client_id", "") or "",
            payload.get("plant_name", "") or "",
            payload.get("plant_id", "") or "",
            payload.get("address", "") or "",
            payload.get("plant_type", "") or "",
            "CON ANOMALIE" if item.get("has_negative") else "POSITIVO",
            " | ".join(ko_labels),
            " | ".join(todo_labels),
            payload.get("general_notes", "") or "",
            item.get("pdf_url", "") or "",
            item.get("server_id", ""),
        ])

    widths = {
        "A": 16, "B": 28, "C": 14, "D": 28, "E": 14, "F": 30,
        "G": 28, "H": 16, "I": 40, "J": 40, "K": 40, "L": 45, "M": 40,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    wb.save(file_path)
    return file_path


def send_email_with_attachment(recipients: List[str], subject: str, body: str, attachment_path: str):
    if not SMTP_HOST or not SMTP_FROM:
        raise HTTPException(status_code=500, detail="SMTP_HOST/SMTP_FROM non configurati")

    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"] = SMTP_FROM
    msg["To"] = ", ".join(recipients)
    msg.set_content(body)

    with open(attachment_path, "rb") as f:
        data = f.read()

    msg.add_attachment(
        data,
        maintype="application",
        subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=os.path.basename(attachment_path),
    )

    if SMTP_PORT == 465:
        with smtplib.SMTP_SSL(SMTP_HOST, SMTP_PORT) as server:
            if SMTP_USER:
                server.login(SMTP_USER, SMTP_PASSWORD)
            server.send_message(msg)
    else:
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as server:
            server.ehlo()
            server.starttls()
            server.ehlo()
            if SMTP_USER:
                server.login(SMTP_USER, SMTP_PASSWORD)
            server.send_message(msg)


@app.post("/checklists")
def create_checklist(payload: ChecklistCreate, request: Request, user=Depends(require_user)):
    server_id = str(uuid.uuid4())

    raw_payload = payload.dict()
    has_ko = any(not c.ok for c in payload.checks)
    has_todo = any(t.selected for t in payload.todos)
    has_negative = has_ko or has_todo

    record = {
        "server_id": server_id,
        "created_by": user["username"],
        "created_at": datetime.utcnow().isoformat(),
        "payload": raw_payload,
        "has_ko": has_ko,
        "has_todo": has_todo,
        "has_negative": has_negative,
        "quote_candidate": has_negative,
        "pdf_created": False,
        "pdf_path": None,
        "pdf_url": None,
        "impianto_pdf_attached": False,
        "syncrogest_upload_response": None,
        "report_sent": False,
        "report_sent_at": None,
    }

    db_insert_checklist(record)

    pdf_path = generate_checklist_pdf(server_id, record)
    pdf_url = build_public_pdf_url(request, server_id)

    impianto_pdf_attached = False
    upload_response = None

    try:
        upload_response = syncrogest_upload_pdf_to_impianto(payload.plant_id, pdf_path)
        impianto_pdf_attached = upload_response["ok"]
    except Exception as e:
        upload_response = {"error": str(e)}

    db_update_checklist_after_pdf(
        server_id=server_id,
        pdf_path=pdf_path,
        pdf_url=pdf_url,
        impianto_pdf_attached=impianto_pdf_attached,
        upload_response=upload_response,
    )

    return {
        "server_id": server_id,
        "quote_candidate": has_negative,
        "pdf_created": True,
        "pdf_url": pdf_url,
        "impianto_pdf_attached": impianto_pdf_attached,
        "syncrogest_upload_response": upload_response,
    }


@app.get("/checklists")
def list_checklists(user=Depends(require_user)):
    return db_list_checklists()


@app.get("/checklists/{server_id}/pdf")
def get_checklist_pdf(server_id: str):
    record = db_get_checklist(server_id)
    if not record:
        raise HTTPException(status_code=404, detail="Checklist non trovata")

    pdf_path = record.get("pdf_path")
    if not pdf_path or not os.path.exists(pdf_path):
        raise HTTPException(status_code=404, detail="PDF non trovato")

    return FileResponse(
        pdf_path,
        media_type="application/pdf",
        filename=os.path.basename(pdf_path),
    )


@app.get("/quote-report/pending")
def quote_report_pending(user=Depends(require_user)):
    return db_list_pending_quote_reports()


@app.post("/quote-report/send")
def quote_report_send(req: QuoteReportSendReq, user=Depends(require_user)):
    items = db_list_pending_quote_reports()

    if not items:
        return {
            "sent": False,
            "message": "Nessuna checklist negativa da inviare",
            "count": 0,
        }

    report_path = create_quote_report_excel(items)

    body = (
        "Buongiorno,\n\n"
        "in allegato il riepilogo delle checklist con esito negativo "
        "da usare per la preparazione dei preventivi.\n\n"
        "Cordiali saluti."
    )

    subject = f"Riepilogo checklist negative - {datetime.now().strftime('%d/%m/%Y %H:%M')}"

    send_email_with_attachment(
        recipients=req.recipients,
        subject=subject,
        body=body,
        attachment_path=report_path,
    )

    if req.mark_sent:
        now_iso = datetime.utcnow().isoformat()
        db_mark_reports_sent([x["server_id"] for x in items], now_iso)

    return {
        "sent": True,
        "count": len(items),
        "report_path": report_path,
        "recipients": req.recipients,
    }


@app.post("/files/upload")
def upload_file(file: UploadFile = File(...), user=Depends(require_user)):
    file_id = str(uuid.uuid4())
    ext = os.path.splitext(file.filename)[1].lower()
    safe_name = f"{file_id}{ext}"
    path = os.path.join(UPLOAD_DIR, safe_name)

    with open(path, "wb") as f:
        f.write(file.file.read())

    db_insert_file(file_id, path)
    return {"file_id": file_id}
